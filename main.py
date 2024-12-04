import openpyxl
from zajecia import Zajecia, Termin
from funkcje import *
import re
from datetime import datetime
import zamiana_danych_na_exel

# Ścieżka do pliku Excel
#plik_excel = 'C:/Users/filip.skup/Desktop/harmonogram.xlsx'
plik_excel = 'harmonogram.xlsx'
#plik_excel = '/Users/filipskup/Desktop/plan_z_xml_do_kalendarza/harmonogram.xlsx'
# Otwórz plik Excel
wb = openpyxl.load_workbook(plik_excel)
wb_termin = openpyxl.load_workbook('harmonogram_tygodnie.xlsx')
# Wybierz arkusz
arkusz = wb['fiz IV']
sheet = wb_termin.active
zajecia_list = []  # lista do przechowywania obiektów zajęć


# Odczytanie nagłówków dni tygodnia (B1-H1)
dni_tygodnia = [sheet.cell(row=1, column=i).value for i in range(2, 9)]  # Zakłada, że dni tygodnia są w kolumnach B-H

# Przechowywanie obiektów Termin w liście
terminy = []

# Iteracja po wierszach (od A2 do A16) z numerem tygodnia
for row in range(2, 17):  # Zakładając, że dane zaczynają się od wiersza 2 i kończą na wierszu 16
    tydzien = sheet.cell(row=row, column=1).value  # Tydzień w kolumnie A
    for col, dzien in enumerate(dni_tygodnia, start=2):  # Kolumny B-H
        data_zajec = sheet.cell(row=row, column=col).value  # Data w danej komórce
        if data_zajec:  # Jeśli komórka zawiera datę
            # Jeżeli data to tekst 'x', to przypisujemy None
            if data_zajec == 'x':
                data_zajec = None
            elif isinstance(data_zajec, str):
                # Konwertujemy na datetime, jeżeli jest to tekst w formacie 'YYYY-MM-DD'
                try:
                    data_zajec = datetime.strptime(data_zajec, "%Y-%m-%d")
                except ValueError:
                    data_zajec = None  # Jeżeli format daty jest niepoprawny, ustawiamy jako None
            termin = Termin(data_zajec, tydzien, dzien)
            terminy.append(termin)

# Wyświetlanie utworzonych terminów
for termin in terminy:
    print(termin)

for kolumna in arkusz.iter_cols(min_row=8, max_row=67, min_col=3, max_col=36):
    for komorka in kolumna:
        if komorka.value is not None:
            komorka_value = komorka.value.strip().lower()
            if re.search(r'\bgr\.\s*11\s*,\s*13\b|\bgr\.\s*13\b|\bgr\s*13\b', komorka_value):
                kolumna_komorki = get_column_letter(komorka)

                dzien = get_dzien_from_kolumna(kolumna_komorki)

                komorka_value = remove_between_zp_gr13(komorka_value)

                komorka_value, tygodnie = split_and_remove_after_gr13(komorka_value)
                godzina_rozpczecia = get_time_start(arkusz, komorka.row)
                godzina_zakonczenia = get_time_end(arkusz, komorka.row, komorka.column)
                # Tworzenie obiektu Zajęcia i dodawanie do listy

                tygodnie_liczby_calkowite = extract_weeks_new(tygodnie)

                zajecie = Zajecia( dzien, komorka.coordinate, komorka_value, tygodnie, godzina_rozpczecia, godzina_zakonczenia)
                zajecie.przypisz_terminy(terminy, tygodnie_liczby_calkowite, zajecie.dzien)
                
                zajecia_list.append(zajecie)


                
for zajecie in zajecia_list:
    print(zajecie)
    print("\nPrzypisane terminy:")
    print(zajecie.wyswietl_terminy())
    print("-" * 40)

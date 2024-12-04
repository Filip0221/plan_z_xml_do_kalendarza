import re
import openpyxl


kolumna_dzien = {
    'C': 'Poniedziałek',
    'D': 'Poniedziałek',
    'E': 'Poniedziałek',
    'F': 'Poniedziałek',
    'G': 'Poniedziałek',
    'H': 'Poniedziałek',
    'I': 'Poniedziałek',
    'J': 'Wtorek',
    'K': 'Wtorek',
    'L': 'Wtorek',
    'M': 'Wtorek',
    'N': 'Wtorek',
    'O': 'Wtorek',
    'P': 'Wtorek',
    'Q': 'Środa',
    'R': 'Środa',
    'S': 'Środa',
    'T': 'Środa',
    'U': 'Środa',
    'V': 'Środa',
    'W': 'Środa',
    'X': 'Czwartek',
    'Y': 'Czwartek',
    'Z': 'Czwartek',
    'AA': 'Czwartek',
    'AB': 'Czwartek',
    'AC': 'Czwartek',
    'AD': 'Piątek',
    'AE': 'Piątek',
    'AF': 'Piątek',
    'AG': 'Piątek',
    'AH': 'Piątek',
    'AI': 'Piątek',
}

def get_dzien_from_kolumna(kolumna_komorki):
    return kolumna_dzien.get(kolumna_komorki, 'Nieznany dzień')

def get_column_letter(komorka):
    coordinate = komorka.coordinate
    if len(coordinate) >= 2 and coordinate[1].isalpha():  # Jeśli kolumna ma dwie litery, np. "AA10"
        return coordinate[:2]  # Zwracamy pierwsze dwie litery
    return coordinate[0]  # Zwracamy pierwszą literę dla kolumn jednoliterowych

def remove_between_zp_gr13(text):
    # Replace everything between (ZP) or (ĆW) and gr. 13 with an empty string
    return re.sub(r'(\(ZP\)|\(ćw\))\s*.*?gr\.\s*13', r'\1 gr. 13', text, flags=re.IGNORECASE)


def split_and_remove_after_gr13(text):
     # Dopasowanie fragmentu tekstu przed i włącznie z gr 13
    text = re.sub(r'13,', '13 ,', text)
    text = re.sub(r',13', ', 13', text)
    match = re.search(r'(.*?13)(.*)', text)
    if match:
        # Pierwsza część: przed gr 13, w tym gr 13 i ewentualnie inne grupy (np. gr 13,14)
        before_gr13 = match.group(1).strip()
        after_gr13 = match.group(2).strip()
        match = re.search(r'\((.*?)\)', after_gr13)
        if match:
            return before_gr13, match.group(1) 
        return before_gr13, after_gr13
    
    return text, text

def get_time_start(arkusz, row):
    godziny_cell = arkusz.cell(row=row, column=2) 
    godzina = godziny_cell.value.strip()
    if godzina:
        match = re.search(r'(.*?)(?:-.*)', godzina)
        return match.group(1).strip()
    return "brak godzin"

def get_time_end(arkusz, row, column):
    komorka = arkusz.cell(row, column)

    for merged_range in arkusz.merged_cells.ranges:
        if komorka.coordinate in merged_range:
            start_col, start_row, end_col, end_row = merged_range.bounds
            godziny_cell = arkusz.cell(row = end_row, column=2)
            godzina = godziny_cell.value.strip()
            if godzina:
                match = re.search(r'-(.*)', godzina)
                return match.group(1).strip()
            return "brak godzin"
        
def przeksztalc_tygodnie_na_liczby(tygodnie):
    # Zamień wszystkie elementy na liczby całkowite, jeśli to możliwe
    tygodnie_liczby = []
    for tydzien in tygodnie:
        try:
            tygodnie_liczby.append(int(tydzien))
        except ValueError:
            # Ignoruj elementy, których nie da się zamienić na liczby całkowite
            continue
    return tygodnie_liczby

def extract_weeks_new( week_range):
        """Funkcja, która przetwarza zakres tygodni"""
        weeks = []
        week_range = week_range.replace(" tydz.", "").replace("tydz.", "")
        parts = week_range.split(',')
        
        for part in parts:
            if '-' in part:
                start, end = part.split('-')
                weeks.extend(range(int(start), int(end) + 1))
            else:
                weeks.append(int(part))
        
        weeks = sorted(set(weeks))
        return weeks
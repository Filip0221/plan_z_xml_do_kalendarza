import openpyxl

import re

kolumna_dzien = {
    'C': 'Poniedziałek',
    'D': 'Poniedziałek',
    'E': 'Poniedziałek',
    'F': 'Poniedziałek',
    'G': 'Poniedziałek',
    'H': 'Poniedziałek',
    'I': 'Poniedziałek',
    'J': 'Wtorek',
    'K': 'Wtorek',
    'L': 'Wtorek',
    'M': 'Wtorek',
    'N': 'Wtorek',
    'O': 'Wtorek',
    'P': 'Wtorek',
    'Q': 'Środa',
    'R': 'Środa',
    'S': 'Środa',
    'T': 'Środa',
    'U': 'Środa',
    'V': 'Środa',
    'W': 'Środa',
    'X': 'Czwartek',
    'Y': 'Czwartek',
    'Z': 'Czwartek',
    'AA': 'Czwartek',
    'AB': 'Czwartek',
    'AC': 'Czwartek',
    'AD': 'Piątek',
    'AE': 'Piątek',
    'AF': 'Piątek',
    'AG': 'Piątek',
    'AH': 'Piątek',
    'AI': 'Piątek',
}


class Zajecia:
    def __init__(self, dzien, adres, wartosc, tygodnie):
        self.dzien = dzien
        self.adres = adres
        self.wartosc = wartosc
        self.tygodnie = tygodnie

    def __str__(self):
        return f"Dzień: {self.dzien}, Adres: {self.adres}, Wartość: {self.wartosc},\n Tygodnie: {self.tygodnie}"


# Ścieżka do pliku Excel
#plik_excel = 'C:/Users/filip.skup/Desktop/harmonogram.xlsx'
plik_excel = '/Users/filipskup/Desktop/plan_z_xml_do_kalendarza/harmonogram.xlsx'
# Otwórz plik Excel
wb = openpyxl.load_workbook(plik_excel)

# Wybierz arkusz
arkusz = wb['fiz IV']

zajecia_list = []  # lista do przechowywania obiektów zajęć

def get_column_letter(komorka):
    coordinate = komorka.coordinate
    if len(coordinate) >= 2 and coordinate[1].isalpha():  # Jeśli kolumna ma dwie litery, np. "AA10"
        return coordinate[:2]  # Zwracamy pierwsze dwie litery
    return coordinate[0]  # Zwracamy pierwszą literę dla kolumn jednoliterowych




def extract_weeks(text):
    # Funkcja wyciąga tygodnie z tekstu
    weeks = re.findall(r'(\d+(?:-\d+)?)\s*tydz\.', text)
    return weeks if weeks else ["Brak informacji o tygodniach"]
def extract_weeks_after_group(text):
    # Poprawione wyrażenie regularne
    match = re.search(r'gr\.\s*13\s*.*?(\([\d\s-]+tydz\.\))', text, re.IGNORECASE)
    if match:
        weeks_str = match.group(1)
        return re.findall(r'\d+(?:-\d+)?', weeks_str)  # Wyszukuje poszczególne tygodnie
    return []

def remove_between_zp_gr13(text):
    # Replace everything between (ZP) or (ĆW) and gr. 13 with an empty string
    return re.sub(r'(\(ZP\)|\(ćw\))\s*.*?gr\.\s*13', r'\1 gr. 13', text, flags=re.IGNORECASE)


def split_and_remove_after_gr13(text):
     # Dopasowanie fragmentu tekstu przed i włącznie z gr 13
    match = re.search(r'(.*13)', text)
    if match:
        # Pierwsza część: przed gr 13, w tym gr 13 i ewentualnie inne grupy (np. gr 13,14)
        before_gr13 = match.group(0).strip()
        return before_gr13
    return text



for kolumna in arkusz.iter_cols(min_row=8, max_row=67, min_col=3, max_col=36):
    for komorka in kolumna:
        if komorka.value is not None:
            komorka_value = komorka.value.strip().lower()
            if re.search(r'\bgr\.\s*11\s*,\s*13\b|\bgr\.\s*13\b|\bgr\s*13\b', komorka_value):
                kolumna_komorki = get_column_letter(komorka)
                dzien = kolumna_dzien.get(kolumna_komorki, 'Nieznany dzień')
                
                # Wyciągnięcie tygodni po "gr. 13"
                #tygodnie = extract_weeks_after_group(komorka.value)

                komorka_value = remove_between_zp_gr13(komorka_value)
                # komorka_value bez niczego po gr 13
                komorka_value = split_and_remove_after_gr13(komorka_value)
                tygodnie = []
                # Tworzenie obiektu Zajęcia i dodawanie do listy

                zajecie = Zajecia(dzien, komorka.coordinate, komorka_value, tygodnie)
                zajecia_list.append(zajecie)

                
for zajecie in zajecia_list:
    print(zajecie)
